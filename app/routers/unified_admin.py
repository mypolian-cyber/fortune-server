from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from app.database import get_db
from app.models.unified_payment import UnifiedPayment
from datetime import datetime, timedelta
from pydantic import BaseModel
from typing import Optional
import httpx, os, secrets, csv, io, random, smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

router = APIRouter()
security = HTTPBasic()

ADMIN_PASSWORD  = os.getenv("ADMIN_PASSWORD", "adelante2026")
PORTONE_SECRET  = os.getenv("PORTONE_SECRET_KEY", "")
ADMIN_EMAIL     = os.getenv("ADMIN_EMAIL", "info@adelante-properties.com")
SMTP_HOST       = os.getenv("SMTP_HOST", "smtp.zoho.com")
SMTP_PORT       = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER       = os.getenv("SMTP_USER", "info@adelante-properties.com")
SMTP_PASSWORD   = os.getenv("SMTP_PASSWORD", "")

# OTP 임시 저장소 {otp_code: expires_at}
otp_store: dict = {}

SERVICE_NAMES = {
    "year_full":  "후아모 올해기운 풀버전",
    "life":       "후아모 평생기운",
    "goonghap":   "후아모 궁합",
    "yukim":      "후아모 지금이질문",
    "ziwei_year": "자미두수 해별운세",
    "ziwei_day":  "자미두수 특정일운세",
}

def verify_admin(credentials: HTTPBasicCredentials = Depends(security)):
    if not secrets.compare_digest(credentials.password, ADMIN_PASSWORD):
        raise HTTPException(
            status_code=401,
            detail="비밀번호가 틀렸습니다",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials

# ─────────────────────────────────────────
# OTP 발송
# ─────────────────────────────────────────
class OTPRequest(BaseModel):
    password: str

class OTPVerify(BaseModel):
    password: str
    otp_code: str

@router.post("/otp/send")
async def send_otp(req: OTPRequest):
    if not secrets.compare_digest(req.password, ADMIN_PASSWORD):
        raise HTTPException(status_code=401, detail="비밀번호가 틀렸습니다")

    # 6자리 OTP 생성
    code = str(random.randint(100000, 999999))
    expires_at = datetime.utcnow() + timedelta(minutes=5)
    otp_store.clear()
    otp_store[code] = expires_at

    # 이메일 발송
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"[어드민] 인증코드: {code}"
        msg["From"]    = SMTP_USER
        msg["To"]      = ADMIN_EMAIL

        html = f"""
        <div style="font-family:sans-serif;max-width:400px;margin:0 auto;padding:24px;background:#f7f4ff;border-radius:12px;">
          <h2 style="color:#1a0a2e;margin-bottom:8px;">🔐 통합 어드민 인증코드</h2>
          <p style="color:#666;margin-bottom:20px;">아래 코드를 입력하세요. 5분 후 만료됩니다.</p>
          <div style="background:#fff;border-radius:10px;padding:20px;text-align:center;">
            <span style="font-size:36px;font-weight:700;letter-spacing:8px;color:#7c3aed;">{code}</span>
          </div>
          <p style="color:#999;font-size:12px;margin-top:16px;">본인이 요청하지 않았다면 즉시 비밀번호를 변경하세요.</p>
        </div>
        """
        msg.attach(MIMEText(html, "html"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASSWORD)
            s.sendmail(SMTP_USER, ADMIN_EMAIL, msg.as_string())
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"이메일 발송 실패: {str(e)}")

    return {"success": True, "message": f"{ADMIN_EMAIL}로 인증코드를 발송했습니다"}

# ─────────────────────────────────────────
# OTP 검증
# ─────────────────────────────────────────
@router.post("/otp/verify")
async def verify_otp(req: OTPVerify):
    if not secrets.compare_digest(req.password, ADMIN_PASSWORD):
        raise HTTPException(status_code=401, detail="비밀번호가 틀렸습니다")

    expires_at = otp_store.get(req.otp_code)
    if not expires_at:
        raise HTTPException(status_code=400, detail="인증코드가 올바르지 않습니다")
    if datetime.utcnow() > expires_at:
        otp_store.pop(req.otp_code, None)
        raise HTTPException(status_code=400, detail="인증코드가 만료되었습니다")

    otp_store.clear()

    # 세션 토큰 발급 (1시간 유효)
    token = secrets.token_hex(32)
    otp_store[f"session_{token}"] = datetime.utcnow() + timedelta(hours=1)

    return {"success": True, "token": token}

# ─────────────────────────────────────────
# 세션 토큰 검증
# ─────────────────────────────────────────
def verify_session(token: str = Query(...)):
    expires_at = otp_store.get(f"session_{token}")
    if not expires_at:
        raise HTTPException(status_code=401, detail="세션이 없습니다. 다시 로그인하세요.")
    if datetime.utcnow() > expires_at:
        otp_store.pop(f"session_{token}", None)
        raise HTTPException(status_code=401, detail="세션이 만료되었습니다. 다시 로그인하세요.")
    return token

# ─────────────────────────────────────────
# 결제 검색
# ─────────────────────────────────────────
@router.get("/payments/search")
async def search_payments(
    birth_date:   Optional[str] = Query(None),
    gender:       Optional[str] = Query(None),
    card_company: Optional[str] = Query(None),
    card_last4:   Optional[str] = Query(None),
    pay_method:   Optional[str] = Query(None),
    service:      Optional[str] = Query(None),
    status:       Optional[str] = Query(None),
    start:        Optional[str] = Query(None),
    end:          Optional[str] = Query(None),
    token:        Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: HTTPBasicCredentials = Depends(verify_admin)
):
    conditions = []
    if birth_date:   conditions.append(UnifiedPayment.birth_date == birth_date)
    if gender:       conditions.append(UnifiedPayment.gender == gender)
    if card_company: conditions.append(UnifiedPayment.card_company.ilike(f"%{card_company}%"))
    if card_last4:   conditions.append(UnifiedPayment.card_last4 == card_last4)
    if pay_method:   conditions.append(UnifiedPayment.pay_method == pay_method)
    if service:      conditions.append(UnifiedPayment.service == service)
    if status:       conditions.append(UnifiedPayment.status == status)
    if start:        conditions.append(UnifiedPayment.paid_at >= start)
    if end:          conditions.append(UnifiedPayment.paid_at <= end + " 23:59:59")

    query = select(UnifiedPayment)
    if conditions:
        query = query.where(and_(*conditions))
    query = query.order_by(UnifiedPayment.paid_at.desc()).limit(100)

    result = await db.execute(query)
    rows = result.scalars().all()

    return [
        {
            "id":           r.id,
            "service":      r.service,
            "service_type": SERVICE_NAMES.get(r.service_type, r.service_type),
            "order_id":     r.order_id,
            "payment_key":  r.payment_key,
            "birth_date":   r.birth_date,
            "gender":       "남" if r.gender == "M" else "여",
            "pay_method":   r.pay_method,
            "card_company": r.card_company or "-",
            "card_last4":   r.card_last4 or "-",
            "amount":       r.amount,
            "refund_amount":r.refund_amount or 0,
            "status":       r.status,
            "paid_at":      r.paid_at.strftime("%Y-%m-%d %H:%M") if r.paid_at else "-",
            "refunded_at":  r.refunded_at.strftime("%Y-%m-%d %H:%M") if r.refunded_at else "-",
            "refund_method":r.refund_method or "-",
            "refund_reason":r.refund_reason or "-",
            "refund_memo":  r.refund_memo or "-",
        }
        for r in rows
    ]

# ─────────────────────────────────────────
# 환불 처리
# ─────────────────────────────────────────
class RefundRequest(BaseModel):
    payment_id:    str
    refund_method: str          # API / CASH
    refund_reason: str
    refund_memo:   Optional[str] = ""   # 현금환불 시 계좌번호

@router.post("/payments/refund")
async def refund_payment(
    req: RefundRequest,
    db: AsyncSession = Depends(get_db),
    _: HTTPBasicCredentials = Depends(verify_admin)
):
    # unified_payments 조회
    result = await db.execute(
        select(UnifiedPayment).where(UnifiedPayment.payment_key == req.payment_id)
    )
    payment = result.scalar_one_or_none()
    if not payment:
        raise HTTPException(status_code=404, detail="결제 내역을 찾을 수 없습니다")
    if payment.status == "refunded":
        raise HTTPException(status_code=400, detail="이미 환불된 결제입니다")

    if req.refund_method == "API":
        # 포트원 환불 API 호출
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                f"https://api.portone.io/payments/{req.payment_id}/cancel",
                headers={"Authorization": f"PortOne {PORTONE_SECRET}"},
                json={"reason": req.refund_reason}
            )
        if r.status_code not in (200, 201):
            raise HTTPException(status_code=502, detail=f"포트원 환불 실패: {r.text}")

    # DB 업데이트
    payment.status        = "refunded"
    payment.refund_amount = payment.amount
    payment.refund_method = req.refund_method
    payment.refund_reason = req.refund_reason
    payment.refund_memo   = req.refund_memo
    payment.refunded_at   = datetime.utcnow()
    await db.commit()

    return {
        "success":      True,
        "order_id":     payment.order_id,
        "refund_amount":payment.amount,
        "refunded_at":  payment.refunded_at.strftime("%Y-%m-%d %H:%M"),
    }

# ─────────────────────────────────────────
# 환불 내역 조회
# ─────────────────────────────────────────
@router.get("/refunds")
async def get_refunds(
    start: Optional[str] = Query(None),
    end:   Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: HTTPBasicCredentials = Depends(verify_admin)
):
    conditions = [UnifiedPayment.status == "refunded"]
    if start: conditions.append(UnifiedPayment.refunded_at >= start)
    if end:   conditions.append(UnifiedPayment.refunded_at <= end + " 23:59:59")

    query = select(UnifiedPayment).where(and_(*conditions)).order_by(UnifiedPayment.refunded_at.desc())
    result = await db.execute(query)
    rows = result.scalars().all()

    total_amount = sum(r.refund_amount or 0 for r in rows)

    return {
        "total_count":  len(rows),
        "total_amount": total_amount,
        "items": [
            {
                "id":           r.id,
                "service":      r.service,
                "service_type": SERVICE_NAMES.get(r.service_type, r.service_type),
                "order_id":     r.order_id,
                "birth_date":   r.birth_date,
                "gender":       "남" if r.gender == "M" else "여",
                "pay_method":   r.pay_method,
                "card_company": r.card_company or "-",
                "card_last4":   r.card_last4 or "-",
                "amount":       r.amount,
                "refund_amount":r.refund_amount,
                "refund_method":r.refund_method,
                "refund_reason":r.refund_reason or "-",
                "refund_memo":  r.refund_memo or "-",
                "paid_at":      r.paid_at.strftime("%Y-%m-%d %H:%M") if r.paid_at else "-",
                "refunded_at":  r.refunded_at.strftime("%Y-%m-%d %H:%M") if r.refunded_at else "-",
            }
            for r in rows
        ]
    }

# ─────────────────────────────────────────
# 환불 내역 엑셀 다운로드 (세무용)
# ─────────────────────────────────────────
@router.get("/refunds/download")
async def download_refunds(
    start: Optional[str] = Query(None),
    end:   Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: HTTPBasicCredentials = Depends(verify_admin)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    conditions = [UnifiedPayment.status == "refunded"]
    if start: conditions.append(UnifiedPayment.refunded_at >= start)
    if end:   conditions.append(UnifiedPayment.refunded_at <= end + " 23:59:59")

    query = select(UnifiedPayment).where(and_(*conditions)).order_by(UnifiedPayment.refunded_at.desc())
    result = await db.execute(query)
    rows = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "환불내역"

    # 스타일 정의
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="3B1A7E")
    card_fill      = PatternFill("solid", fgColor="EDE9FE")
    cash_fill      = PatternFill("solid", fgColor="FEF3C7")
    total_font     = Font(bold=True, size=11)
    total_fill     = PatternFill("solid", fgColor="F3F0FF")
    center         = Alignment(horizontal="center", vertical="center")
    thin           = Side(style="thin", color="E5E5E5")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)
    amount_fill    = PatternFill("solid", fgColor="FEE2E2")

    # 제목행
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = f"환불 내역 ({start or '전체'} ~ {end or '전체'})"
    title_cell.font  = Font(bold=True, size=13, color="1A0A2E")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    # 헤더
    headers = [
        "번호", "서비스", "주문번호", "생년월일", "성별",
        "결제수단", "카드사", "카드뒷4자리",
        "결제금액(원)", "환불금액(원)", "결제일시", "환불일시",
        "환불형태", "환불사유", "메모"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[2].height = 22

    # 데이터행
    total_amount = 0
    total_refund = 0
    for i, r in enumerate(rows, 1):
        row_num = i + 2
        is_cash = r.refund_method == "CASH"
        row_fill = cash_fill if is_cash else card_fill

        values = [
            i,
            "후아모" if r.service == "huamo" else "자미두수",
            r.order_id or "-",
            r.birth_date or "-",
            "남" if r.gender == "M" else "여",
            r.pay_method or "-",
            r.card_company or "-",
            r.card_last4 or "-",
            r.amount,
            r.refund_amount or 0,
            r.paid_at.strftime("%Y-%m-%d %H:%M") if r.paid_at else "-",
            r.refunded_at.strftime("%Y-%m-%d %H:%M") if r.refunded_at else "-",
            "💵 현금환불" if is_cash else "💳 카드환불(PG)",
            r.refund_reason or "-",
            r.refund_memo or "-",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border    = border
            cell.alignment = center
            # 금액 컬럼 색상
            if col == 9:
                cell.fill = PatternFill("solid", fgColor="DBEAFE")
            elif col == 10:
                cell.fill = amount_fill
            else:
                cell.fill = row_fill

        total_amount += r.amount
        total_refund += r.refund_amount or 0

    # 합계행
    total_row = len(rows) + 3
    ws.cell(row=total_row, column=1, value="합계").font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).alignment = center
    for col in range(2, 9):
        ws.cell(row=total_row, column=col).fill = total_fill
    amt_cell = ws.cell(row=total_row, column=9,  value=total_amount)
    ref_cell = ws.cell(row=total_row, column=10, value=total_refund)
    amt_cell.font = Font(bold=True, color="1E40AF")
    ref_cell.font = Font(bold=True, color="991B1B")
    amt_cell.fill = PatternFill("solid", fgColor="DBEAFE")
    ref_cell.fill = amount_fill
    amt_cell.alignment = center
    ref_cell.alignment = center
    for col in range(11, 16):
        ws.cell(row=total_row, column=col).fill = total_fill

    # 컬럼 너비
    col_widths = [6, 10, 30, 14, 6, 12, 12, 12, 14, 14, 18, 18, 16, 24, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 범례 시트
    ws2 = wb.create_sheet("범례")
    ws2["A1"] = "환불형태 구분"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2["A3"] = "💳 카드환불(PG)"
    ws2["B3"] = "포트원 API를 통한 카드 자동취소 → 부가세 매출취소 처리"
    ws2["A3"].fill = card_fill
    ws2["A4"] = "💵 현금환불"
    ws2["B4"] = "계좌이체로 직접 송금한 환불 → 현금지출로 별도 관리"
    ws2["A4"].fill = cash_fill
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 50

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)

    period = f"{start or 'all'}_{end or 'all'}"
    filename = f"refunds_{period}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ─────────────────────────────────────────
# 통계 API (기간별 이용자 분석)
# ─────────────────────────────────────────
@router.get("/stats/users")
async def get_user_stats(
    start: Optional[str] = Query(None),
    end:   Optional[str] = Query(None),
    token: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: HTTPBasicCredentials = Depends(verify_admin)
):
    import psycopg2

    now = datetime.now()

    # fortune DB (saju_cache) 조회
    fortune_rows = []
    try:
        conn = psycopg2.connect(
            "postgresql://fortune_user:fortune_pass_2026@localhost:5432/fortune"
        )
        cur = conn.cursor()
        query = "SELECT year, gender, created_at FROM saju_cache WHERE 1=1"
        params = []
        if start:
            query += " AND created_at >= %s"
            params.append(start)
        if end:
            query += " AND created_at <= %s"
            params.append(end + " 23:59:59")
        cur.execute(query, params)
        fortune_rows = [{"year": r[0], "gender": r[1], "created_at": r[2], "service": "huamo"} for r in cur.fetchall()]
        cur.close()
        conn.close()
    except Exception as e:
        print(f"fortune DB 조회 오류: {e}")

    # ziwei DB (ziwei_cache) 조회
    ziwei_rows = []
    try:
        conn = psycopg2.connect(
            "postgresql://fortune_user:fortune_pass_2026@localhost:5432/ziwei"
        )
        cur = conn.cursor()
        query = "SELECT input_data, created_at FROM ziwei_cache WHERE 1=1"
        params = []
        if start:
            query += " AND created_at >= %s"
            params.append(start)
        if end:
            query += " AND created_at <= %s"
            params.append(end + " 23:59:59")
        cur.execute(query, params)
        import json as _json
        for r in cur.fetchall():
            try:
                d = _json.loads(r[0]) if r[0] else {}
                is_male = d.get("is_male")
                gender = "M" if is_male is True else ("F" if is_male is False else None)
                ziwei_rows.append({"year": d.get("year"), "gender": gender, "created_at": r[1], "service": "ziwei"})
            except:
                pass
        cur.close()
        conn.close()
    except Exception as e:
        print(f"ziwei DB 조회 오류: {e}")

    all_rows = fortune_rows + ziwei_rows

    # 성별 집계
    gender_data = {"M": 0, "F": 0, "unknown": 0}
    for r in all_rows:
        g = r.get("gender")
        if g == "M":   gender_data["M"] += 1
        elif g == "F": gender_data["F"] += 1
        else:          gender_data["unknown"] += 1

    # 연령대 집계 (성별 포함)
    age_keys = ["10대", "20대", "30대", "40대", "50대", "60대+"]
    age_groups = {k: {"M": 0, "F": 0, "total": 0} for k in age_keys}
    for r in all_rows:
        birth_year = r.get("year")
        if birth_year:
            age = now.year - birth_year
            if age < 20:   key = "10대"
            elif age < 30: key = "20대"
            elif age < 40: key = "30대"
            elif age < 50: key = "40대"
            elif age < 60: key = "50대"
            else:          key = "60대+"
            g = r.get("gender")
            if g == "M":   age_groups[key]["M"] += 1
            elif g == "F": age_groups[key]["F"] += 1
            age_groups[key]["total"] += 1

    # 일자별 이용자 (서비스별 분리)
    daily = {}
    for r in all_rows:
        dt = r.get("created_at")
        if dt:
            day = dt.strftime("%Y-%m-%d") if hasattr(dt, 'strftime') else str(dt)[:10]
            if day not in daily:
                daily[day] = {"huamo": 0, "ziwei": 0}
            svc = r.get("service", "")
            if svc == "huamo":
                daily[day]["huamo"] += 1
            elif svc == "ziwei":
                daily[day]["ziwei"] += 1

    # 서비스별
    service_data = {"huamo": 0, "ziwei": 0}
    for r in all_rows:
        s = r.get("service")
        if s in service_data:
            service_data[s] += 1

    return {
        "total": len(all_rows),
        "gender": gender_data,
        "age_groups": age_groups,
        "daily": dict(sorted(daily.items())),
        "services": service_data,
    }

# ─────────────────────────────────────────
# 통계 엑셀 다운로드
# ─────────────────────────────────────────
@router.get("/stats/download")
async def download_stats(
    start: Optional[str] = Query(None),
    end:   Optional[str] = Query(None),
    token: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: HTTPBasicCredentials = Depends(verify_admin)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import psycopg2, json as _json, io as _io

    now = datetime.now()

    # 데이터 수집 (위와 동일)
    all_rows = []
    for db_name, service_name in [("fortune", "huamo"), ("ziwei", "ziwei")]:
        try:
            conn = psycopg2.connect(f"postgresql://fortune_user:fortune_pass_2026@localhost:5432/{db_name}")
            cur = conn.cursor()
            if service_name == "huamo":
                q = "SELECT year, gender, created_at FROM saju_cache WHERE 1=1"
            else:
                q = "SELECT input_data, created_at FROM ziwei_cache WHERE 1=1"
            params = []
            if start:
                q += " AND created_at >= %s"; params.append(start)
            if end:
                q += " AND created_at <= %s"; params.append(end + " 23:59:59")
            cur.execute(q, params)
            for r in cur.fetchall():
                if service_name == "huamo":
                    all_rows.append({"year": r[0], "gender": r[1], "created_at": r[2], "service": "후아모"})
                else:
                    try:
                        d = _json.loads(r[0]) if r[0] else {}
                        is_male = d.get("is_male")
                        g = "M" if is_male is True else ("F" if is_male is False else None)
                        all_rows.append({"year": d.get("year"), "gender": g, "created_at": r[1], "service": "자미두수"})
                    except: pass
            cur.close(); conn.close()
        except Exception as e:
            print(f"{db_name} 오류: {e}")

    wb = Workbook()

    # 스타일
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="3B1A7E")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="E5E5E5")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 시트1: 일자별 이용자 ──
    ws1 = wb.active
    ws1.title = "일자별이용자"
    ws1.merge_cells("A1:D1")
    ws1["A1"].value = f"일자별 이용자 현황 ({start or '전체'} ~ {end or '전체'})"
    ws1["A1"].font  = Font(bold=True, size=13, color="1A0A2E")
    ws1["A1"].alignment = center
    ws1.row_dimensions[1].height = 28

    for col, h in enumerate(["날짜", "후아모", "자미두수", "합계"], 1):
        c = ws1.cell(row=2, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = border

    daily_huamo  = {}
    daily_ziwei  = {}
    for r in all_rows:
        dt = r.get("created_at")
        day = dt.strftime("%Y-%m-%d") if hasattr(dt, 'strftime') else str(dt)[:10]
        if r["service"] == "후아모":
            daily_huamo[day] = daily_huamo.get(day, 0) + 1
        else:
            daily_ziwei[day] = daily_ziwei.get(day, 0) + 1

    all_days = sorted(set(list(daily_huamo.keys()) + list(daily_ziwei.keys())))
    total_h = total_z = 0
    for i, day in enumerate(all_days, 3):
        h = daily_huamo.get(day, 0)
        z = daily_ziwei.get(day, 0)
        total_h += h; total_z += z
        for col, val in enumerate([day, h, z, h+z], 1):
            c = ws1.cell(row=i, column=col, value=val)
            c.alignment = center; c.border = border
            if col == 4: c.font = Font(bold=True)

    tr = len(all_days) + 3
    for col, val in enumerate(["합계", total_h, total_z, total_h+total_z], 1):
        c = ws1.cell(row=tr, column=col, value=val)
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="F3F0FF")
        c.alignment = center; c.border = border

    for col, w in enumerate([14, 10, 10, 10], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── 시트2: 성별/연령 분석 ──
    ws2 = wb.create_sheet("성별연령분석")
    ws2["A1"].value = "성별 분포"
    ws2["A1"].font  = Font(bold=True, size=12, color="1A0A2E")

    gender_data = {"남성": 0, "여성": 0}
    age_groups  = {"10대": 0, "20대": 0, "30대": 0, "40대": 0, "50대": 0, "60대+": 0}
    for r in all_rows:
        g = r.get("gender")
        if g == "M": gender_data["남성"] += 1
        elif g == "F": gender_data["여성"] += 1
        birth_year = r.get("year")
        if birth_year:
            age = now.year - birth_year
            if age < 20:   age_groups["10대"] += 1
            elif age < 30: age_groups["20대"] += 1
            elif age < 40: age_groups["30대"] += 1
            elif age < 50: age_groups["40대"] += 1
            elif age < 60: age_groups["50대"] += 1
            else:          age_groups["60대+"] += 1

    for col, h in enumerate(["성별", "인원수", "비율"], 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = border

    total_gender = sum(gender_data.values()) or 1
    fills = {"남성": "DBEAFE", "여성": "FCE7F3"}
    for i, (g, cnt) in enumerate(gender_data.items(), 3):
        pct = f"{cnt/total_gender*100:.1f}%"
        for col, val in enumerate([g, cnt, pct], 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=fills.get(g, "FFFFFF"))
            c.alignment = center; c.border = border

    ws2["A7"].value = "연령대 분포"
    ws2["A7"].font  = Font(bold=True, size=12, color="1A0A2E")

    for col, h in enumerate(["연령대", "인원수", "비율"], 1):
        c = ws2.cell(row=8, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = border

    total_age = sum(age_groups.values()) or 1
    age_fills = ["EDE9FE","DDD6FE","C4B5FD","A78BFA","8B5CF6","7C3AED"]
    for i, (age, cnt) in enumerate(age_groups.items(), 9):
        pct = f"{cnt/total_age*100:.1f}%"
        for col, val in enumerate([age, cnt, pct], 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=age_fills[i-9])
            c.alignment = center; c.border = border

    for col, w in enumerate([12, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── 시트3: 서비스별 분석 ──
    ws3 = wb.create_sheet("서비스별분석")
    ws3["A1"].value = "서비스별 이용 현황"
    ws3["A1"].font  = Font(bold=True, size=12, color="1A0A2E")

    for col, h in enumerate(["서비스", "이용자수", "비율"], 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = border

    svc_data = {"후아모": 0, "자미두수": 0}
    for r in all_rows:
        s = r.get("service")
        if s in svc_data: svc_data[s] += 1

    total_svc = sum(svc_data.values()) or 1
    svc_fills = {"후아모": "DBEAFE", "자미두수": "FCE7F3"}
    for i, (svc, cnt) in enumerate(svc_data.items(), 3):
        pct = f"{cnt/total_svc*100:.1f}%"
        for col, val in enumerate([svc, cnt, pct], 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=svc_fills.get(svc, "FFFFFF"))
            c.alignment = center; c.border = border

    for col, w in enumerate([14, 10, 10], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)

    period = f"{start or 'all'}_{end or 'all'}"
    filename = f"user_stats_{period}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
